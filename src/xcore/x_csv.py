# -*- coding: UTF-8 -*-
from x_log import x_log
from x_utf8 import x_utf8
# from openpyxl import load_workbook
from openpyxl import *
import os


class x_csv:
    _inpath = ""
    _workbook = None
    _err = None
    _isCreateFile = False

    def __init__(self, inpath):
        self._inpath = inpath
        pass

    def _Check(self):
        if self._workbook == None:
            raise "not call load or Create"

    def Load(self):
        if os.path.isfile(self._inpath) == True:
            # self._workbook = load_workbook(self._inpath)
            # self._workbook = load_workbook(self._inpath,keep_vba=True,data_only=True); #不能保存vba
            self._workbook = load_workbook(
                self._inpath, keep_vba=True, data_only=False)  # 能保存vba
        else:
            x_log.Error("workbook not exist!")

    # 只处理csv 和 xlsx
    def Create(self, forceCreate: bool = False):
        if os.path.isfile(self._inpath) == True:
            if(forceCreate):
                os.remove(self._inpath)
                self._workbook = Workbook()  # 在save的时候创建
                self._isCre_isCreateFileateFil = True
            else:
                x_log.Error("workbook exist!")
        else:
            self._workbook = Workbook()  # 在save的时候创建
            self._isCre_isCreateFileateFil = True

    # def reload(self):
    #     self._workbook = load_workbook(self._inpath)
    #     pass

    def Save(self):
        # if(self._isCreateFile):
        #     x_log.Info("创建:" + self._inpath)
        # else:
        #     x_log.Info("保存:" + self._inpath)
        pathslpit = os.path.split(self._inpath)
        if not os.path.exists(pathslpit[0]):
            os.mkdir(pathslpit[0])
        self._workbook.save(self._inpath)

        pass

    def ModifyCell(self, wsname: str, cellName: str, value):
        self._Check()
        if self._err != None:
            x_log.Error(self._err)
        ws = self._workbook[wsname]
        ws[cellName] = value
        pass

    def AddSheet(self, sheetname: str):
        ws = self._workbook.create_sheet(sheetname, 0)
        ws.Title = sheetname

    # def UseSheet(self, sheetname: str):
    #     ws = self._workbook[sheetname]
    #     ws.Title = sheetname

    def EachRows(self, worksheetname, action):
        self._Check()
        # columns = ws.max_column;
        ws = self._workbook[worksheetname]
        maxrow = ws.max_row
        maxcolumn = ws.max_column
        for y in range(1, maxrow + 1):
            ret = []
            for x in range(1, maxcolumn + 1):
                ret.append(ws.cell(row=y, column=x).value)
                # x_log.Info(ws.cell(row=y, column=x).value)
            action(ret)

    def AllRows(self, worksheetname) -> any:
        self._Check()
        rows = []
        def f(x): rows.append(x)
        self.EachRows(worksheetname, f)
        return rows

    def SetValue(self, sheet: str, row: int, column: int, value: any):
        self._Check()
        ws = self._workbook[sheet]
        ws.cell(row, column, value)

    @property
    def Path(self) ->str:
        return self._inpath;

    @property
    def Sheets(self)->list:
        self._Check()
        return self._workbook

        
    @property
    def SheetNames(self)->list:
        self._Check()
        return self._workbook.sheetnames